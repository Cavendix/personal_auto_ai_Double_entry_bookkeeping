"""
write_ledger.py — Skill: 写入账本
读取 ocr_done.csv 中 status=success 且 written_to_ledger=false 的行，
解析 AI 返回的 JSON，支持单笔和多笔拆分，追加写入 ledger.xlsx。
"""
import json
import sys
from datetime import datetime

import openpyxl
from dotenv import load_dotenv

sys.path.insert(0, str(__import__("pathlib").Path(__file__).parent))
import utils

load_dotenv()

log = utils.get_logger("write_ledger")

# 账本表头，与 README 字段定义一致
LEDGER_HEADERS = [
    "time", "type", "amount", "source", "dest",
    "note", "raw_input", "image_path", "processed_at",
]


def _get_or_create_sheet(wb: openpyxl.Workbook):
    ws = wb.active
    # 如果第一行为空说明是新建文件，写表头
    if ws.max_row == 0 or ws.cell(1, 1).value is None:
        ws.append(LEDGER_HEADERS)
    return ws


def _parse_entries(ai_result: str, row: dict) -> list[dict]:
    """
    解析 AI 返回的 JSON，统一转为 list[dict]。
    支持单笔 {...} 和多笔 [{...}, {...}]。
    """
    data = json.loads(ai_result)
    if isinstance(data, dict):
        entries = [data]
    elif isinstance(data, list):
        entries = data
    else:
        raise ValueError(f"AI 返回格式异常: {type(data)}")

    result = []
    for entry in entries:
        if "error" in entry:
            log.warning(f"AI 标记无法识别: {entry['error']} — {row['image_path']}")
            continue
        result.append({
            "time":         entry.get("date") or "",
            "type":         entry.get("type") or "",
            "amount":       entry.get("amount") or 0,
            "source":       entry.get("source") or "",
            "dest":         entry.get("dest") or "",
            "note":         entry.get("note") or "",
            "raw_input":    row["ocr_text"],
            "image_path":   row["image_path"],
            "processed_at": utils.now_str(),
        })
    return result


def run() -> str:
    done_rows = utils.read_csv(utils.OCR_DONE, utils.OCR_DONE_FIELDS)
    pending   = [r for r in done_rows
                 if r.get("status") == "success" and r.get("written_to_ledger") == "false"]

    if not pending:
        msg = "ocr_done.csv 中没有待入账的成功记录。"
        log.info(msg)
        return msg

    # 打开或新建 Excel
    if utils.LEDGER_PATH.exists():
        wb = openpyxl.load_workbook(utils.LEDGER_PATH)
    else:
        wb = openpyxl.Workbook()

    ws = _get_or_create_sheet(wb)

    written, skipped = 0, 0

    for row in pending:
        try:
            entries = _parse_entries(row["ai_result"], row)
            for entry in entries:
                ws.append([entry[h] for h in LEDGER_HEADERS])
                written += 1
            row["written_to_ledger"] = "true"
            log.info(f"入账 {len(entries)} 笔: {row['image_path']}")
        except Exception as e:
            skipped += 1
            log.error(f"解析失败跳过: {row['image_path']} — {e}")

    wb.save(utils.LEDGER_PATH)

    # 更新 ocr_done.csv 中的 written_to_ledger 标记
    utils.rewrite_csv(utils.OCR_DONE, done_rows, utils.OCR_DONE_FIELDS)

    msg = f"入账完成：写入 {written} 笔，跳过 {skipped} 条（解析失败）。"
    log.info(msg)
    return msg


if __name__ == "__main__":
    print(run())
