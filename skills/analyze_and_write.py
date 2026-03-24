"""
analyze_and_write.py — Skill: 批量 AI 分析与直接入账
读取 ocr_queue.csv 中未分析的行，调用 AI 提取记账 JSON，
提取字段（包含余额），直接追加写入 ledger.xlsx，并清空队列。
"""
import json
import os
import subprocess
import sys
import tempfile
import openpyxl

import yaml
from dotenv import load_dotenv

# 将上级目录加入 sys.path，以便导入 utils 模块
sys.path.insert(0, str(__import__("pathlib").Path(__file__).parent))
import utils

load_dotenv()

# 从环境变量获取配置，默认使用 cli 模式调用 gemini
ANALYZE_MODE     = os.getenv("ANALYZE_MODE", "cli").lower()
GEMINI_CMD       = os.getenv("GEMINI_CLI_CMD", "gemini")
GEMINI_API_KEY   = os.getenv("GEMINI_API_KEY", "")
GEMINI_API_MODEL = os.getenv("GEMINI_API_MODEL", "gemini-2.0-flash")
ANALYZE_TIMEOUT  = int(os.getenv("ANALYZE_TIMEOUT", "60"))

log = utils.get_logger("analyze_and_write")

# 定义最终写入 Excel 的表头顺序
LEDGER_HEADERS = [
    "time", "type", "amount", "source", "dest", "balance",
    "note", "image_path",
]

def _load_prompt() -> tuple[str, str]:
    """读取 prompts.yaml，返回元组: (系统提示词, 用户提示词模板)。"""
    with open(utils.PROMPTS, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    return data["system"], data["user_template"]

def _call_gemini_cli(system: str, user_msg: str) -> str:
    """
    方式 A：调用本地安装的 gemini-cli。
    为了避免命令行长度限制，将提示词写入临时文件并使用 @file 方式传递。
    """
    prompt = f"{system}\n\n---\n\n{user_msg}"
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".txt", delete=False, encoding="utf-8"
    ) as tmp:
        tmp.write(prompt)
        tmp_path = tmp.name
    try:
        # 调用命令行，捕获输出
        result = subprocess.run(
            [GEMINI_CMD, "-p", f"@{tmp_path}"],
            capture_output=True,
            text=True,
            timeout=ANALYZE_TIMEOUT,
            encoding="utf-8",
        )
        if result.returncode != 0:
            raise RuntimeError(result.stderr.strip() or "gemini-cli returned non-zero code")
        return result.stdout.strip()
    finally:
        os.unlink(tmp_path) # 执行完毕后删除临时文件

def _call_gemini_api(system: str, user_msg: str) -> str:
    """
    方式 B：调用 Google Gemini HTTP API。
    需要在 .env 中配置 GEMINI_API_KEY 并在环境中安装 google-generativeai。
    """
    try:
        import google.generativeai as genai
    except ImportError:
        raise ImportError("pip install google-generativeai is required for api mode")
    if not GEMINI_API_KEY:
        raise ValueError("GEMINI_API_KEY is not set")
    
    # 初始化并调用模型
    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel(
        model_name=GEMINI_API_MODEL,
        system_instruction=system,
    )
    response = model.generate_content(user_msg)
    return response.text.strip()

def _call_ai(system: str, user_msg: str) -> str:
    """根据环境变量 ANALYZE_MODE 选择合适的 AI 调用方式"""
    if ANALYZE_MODE == "api":
        log.debug("Using Gemini API mode")
        return _call_gemini_api(system, user_msg)
    log.debug("Using gemini-cli mode")
    return _call_gemini_cli(system, user_msg)

def _extract_json(raw: str) -> str:
    """
    清洗 AI 返回的结果，剥离 Markdown 代码块标记（如 ```json 和 ```），
    提取出纯净的 JSON 字符串。
    """
    raw = raw.strip()
    if raw.startswith("```"):
        lines = [l for l in raw.splitlines() if not l.startswith("```")]
        raw = "\n".join(lines).strip()
    return raw

def _get_or_create_sheet(wb: openpyxl.Workbook):
    """
    检查或初始化 Excel 表格。如果文件为空，则写入标准表头。
    会对齐当前定义的 LEDGER_HEADERS，清理过时的列，或者在需要时插入新列（如 balance 列）。
    """
    ws = wb.active
    # 如果是空表，直接追加表头
    if ws.max_row == 0 or ws.cell(1, 1).value is None:
        ws.append(LEDGER_HEADERS)
    else:
        headers = [cell.value for cell in ws[1]]
        # 移除旧版代码留下的弃用列
        for target in ["processed_at", "raw_input"]:
            if target in headers:
                idx = headers.index(target) + 1
                ws.delete_cols(idx)
                headers.pop(idx - 1)
        
        current_headers = [cell.value for cell in ws[1]]
        # 如果旧账本缺少 balance 列，自动在指定位置插入
        if "balance" not in current_headers:
            if "note" in current_headers and "image_path" in current_headers:
                ws.insert_cols(6)  # 插入在第 6 列（即 'F' 列）
            # 重新覆写第一行，以确保与 LEDGER_HEADERS 严格对齐
            for col_idx, header in enumerate(LEDGER_HEADERS, start=1):
                ws.cell(row=1, column=col_idx).value = header
    return ws

def _parse_entries(cleaned: str, row: dict) -> list[dict]:
    """
    解析 AI 提取的 JSON 数据。
    支持返回单笔交易字面量（dict）或者多笔交易（list）。
    同时对解析出的 amount 和 balance 字段进行数字格式清洗。
    """
    data = json.loads(cleaned)
    if isinstance(data, dict):
        entries = [data]
    elif isinstance(data, list):
        entries = data
    else:
        raise ValueError(f"Abnormal AI response format: {type(data)}")

    result = []
    for entry in entries:
        if "error" in entry:
            raise ValueError(f"AI error: {entry['error']}")
        
        # 清洗余额（balance）
        balance_val = entry.get("balance")
        if balance_val == "null" or balance_val is None or balance_val == "":
            balance_val = ""
        else:
            try:
                balance_val = float(balance_val)
            except ValueError:
                balance_val = ""

        # 清洗金额（amount）
        amount_val = entry.get("amount", 0)
        try:
            amount_val = float(amount_val)
        except ValueError:
            amount_val = 0.0

        # 返回与表头一一对应的字典数据
        result.append({
            "time":         entry.get("date") or "",
            "type":         entry.get("type") or "",
            "amount":       amount_val,
            "source":       entry.get("source") or "",
            "dest":         entry.get("dest") or "",
            "balance":      balance_val,
            "note":         entry.get("note") or "",
            "image_path":   row["image_path"],
        })
    return result

def run() -> str:
    """
    主执行函数：读取待处理队列，利用 AI 进行结构化提取，处理后直接追加写入 Excel 账本，最后清空分析队列。
    供 MCP Server 作为 tool 使用，也会返回相应的文字播报。
    """
    system, user_tpl = _load_prompt()

    # 从 ocr_queue 读取待分析的数据
    queue = utils.read_csv(utils.OCR_QUEUE, utils.OCR_QUEUE_FIELDS)
    if not queue:
        msg = "ocr_queue.csv is empty. No new transactions."
        log.info(msg)
        return msg

    # 加载或创建 Excel 账本
    if utils.LEDGER_PATH.exists():
         wb = openpyxl.load_workbook(utils.LEDGER_PATH)
    else:
         wb = openpyxl.Workbook()
    ws = _get_or_create_sheet(wb)

    success, failed = 0, 0

    # 逐行处理队列中的识别结果
    for row in queue:
        image_path = row["image_path"]
        ocr_text   = row["ocr_text"]
        log.info(f"AI Analyzing & Writing: {image_path}")

        user_msg = user_tpl.replace("{ocr_text}", ocr_text)

        try:
            raw     = _call_ai(system, user_msg)        # 获取 AI 生文字响应
            cleaned = _extract_json(raw)                # 提取 JSON 对象
            entries = _parse_entries(cleaned, row)      # 转换处理为 Python 字典列表

            # 将条目追加进表格末尾
            for entry in entries:
                ws.append([entry[h] for h in LEDGER_HEADERS])
            
            success += len(entries)
            log.info(f"Success: written {len(entries)} entries for {image_path}")

        except Exception as e:
            failed += 1
            log.error(f"Failed to process {image_path} — {e}")

    # 保存账本，并清空队列供下次使用
    wb.save(utils.LEDGER_PATH)
    utils.rewrite_csv(utils.OCR_QUEUE, [], utils.OCR_QUEUE_FIELDS)

    msg = f"Analysis & Ledger update complete: {success} entries successfully written, {failed} files failed. Details in logs/app.log."
    log.info(msg)
    return msg

if __name__ == "__main__":
    print(run())
