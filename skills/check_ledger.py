"""
check_ledger.py — Skill: 复式记账检查
根据 ledger.xlsx 中的支出和收入计算每个账户的变动，并与 OCR 提取的余额列对比。
如发现不一致，则报告具体日期、行号及误差情况。
"""
import sys
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

# 将上级目录加入 sys.path，以便导入 utils 模块
sys.path.insert(0, str(Path(__file__).parent))
import utils

load_dotenv()

log = utils.get_logger("check_ledger")

def run() -> str:
    """
    检查账本的自洽性。
    逐行模拟交易并推算账本内各账户在当下的理论余额。
    当遇到具有确切识别余额（balance列）的行时，与推算余额进行比对，
    记录任何由于误差、漏记引起的余额不匹配，供使用者排查。
    """
    # 确认账本文件是否存在
    if not utils.LEDGER_PATH.exists():
        msg = "ledger.xlsx 不存在，无法进行检查。"
        log.info(msg)
        return msg

    # 读取表格记录（仅加载数据值）
    wb = openpyxl.load_workbook(utils.LEDGER_PATH, data_only=True)
    ws = wb.active

    if ws.max_row <= 1:
        msg = "ledger.xlsx 中没有数据行。"
        log.info(msg)
        return msg

    # 动态获取表头所在的列索引，解耦对于列位置的强依赖
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    try:
        time_idx = headers.index("time")
        type_idx = headers.index("type")
        amt_idx  = headers.index("amount")
        src_idx  = headers.index("source")
        dest_idx = headers.index("dest")
        bal_idx  = headers.index("balance")
    except ValueError as e:
        msg = f"ledger.xlsx 表头缺失必需字段，检查失败: {e}"
        log.error(msg)
        return msg

    computed_balances = {}       # 存放我们在代码推演中得出的每个账户余额
    initialized_accounts = set() # 记录哪些账户已经用从表格获得的“初始真实余额”初始化了
    errors = []                  # 收集发现的问题

    # 从第 2 行（跳过表头）开始逐行扫描账本数据
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_time = row[time_idx] or ""
        row_type = str(row[type_idx] or "").strip()
        
        # 解析金额，如果无法解析则视为 0
        try:
            amount = float(row[amt_idx])
        except (ValueError, TypeError):
            amount = 0.0

        source = str(row[src_idx] or "").strip()
        dest = str(row[dest_idx] or "").strip()
        file_balance_raw = row[bal_idx]

        # 根据交易类型，更新相关账户的理论余额
        if row_type == "支出" and source:
            computed_balances[source] = computed_balances.get(source, 0.0) - amount
        elif row_type == "收入" and dest:
            computed_balances[dest] = computed_balances.get(dest, 0.0) + amount
        elif row_type == "互转":
            if source:
                computed_balances[source] = computed_balances.get(source, 0.0) - amount
            if dest:
                computed_balances[dest] = computed_balances.get(dest, 0.0) + amount

        # 判断这笔交易之后，表格里的余额数据（如有）属于哪一方账户
        target_account = ""
        if row_type == "支出":
            target_account = source
        elif row_type in ["收入", "互转"]:
            target_account = dest
        else:
            # 对于未知类型或无归属对象的，跳过检查
            pass
        
        # 开始验证比对
        if target_account and file_balance_raw is not None and str(file_balance_raw).strip() != "":
            try:
                # 尝试解析表中的核对余额
                fb = float(file_balance_raw)
                if target_account not in initialized_accounts:
                    # 如果这是本账户第一次遇到带有余额的记录，
                    # 那么以此作为起点基准值，赋给推算状态。
                    computed_balances[target_account] = fb
                    initialized_accounts.add(target_account)
                else:
                    cb = computed_balances[target_account]
                    # 针对经过数学推算后的余额与表内的提取余额进行浮点差异对比（容差值 0.05 元）
                    if abs(cb - fb) > 0.05:
                        err_msg = (f"行 {row_idx} ({row_time}): 账户[{target_account}] 余额不符。 "
                                   f"计算推算值: {cb:.2f}, 账本原值: {fb:.2f}")
                        errors.append(err_msg)
            except ValueError:
                pass  # 表内的余额非合法浮点数字符串，跳过检查

    # 根据检查结果提供详细反馈
    if errors:
        msg = f"复式记账检查发现 {len(errors)} 个计算与现实值出入的错误:\n" + "\n".join(errors)
        log.warning(msg)
        return msg
    else:
        msg = "复式记账检查通过，未发现余额异常。"
        log.info(msg)
        return msg

if __name__ == "__main__":
    print(run())
